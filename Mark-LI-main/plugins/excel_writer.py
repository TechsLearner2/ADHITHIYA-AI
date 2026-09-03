"""
JARVIS plugin — Excel writer (voice → real .xlsx).

The user describes a spreadsheet in plain speech — "make me a table of my
five expenses this week with a total", "bana 12 aylık satış tablosu ve bir
grafik oluştur" — and JARVIS produces a genuine .xlsx on the Desktop:
bold headers, typed cells, an optional SUM total row, and an optional
chart. Gemini turns the request into a structured spec; openpyxl builds
the file. That last step is the whole point — Gemini can describe a table
but it cannot hand you a working .xlsx; this plugin does.

Cross-OS (saves to the user's Desktop, falling back to home). Needs
openpyxl (`pip install openpyxl`). All languages — the sheet content stays
in whatever language the user asked for.
"""

import json
import re
from datetime import datetime
from pathlib import Path

PLUGIN = {
    "name": "excel_writer",
    "description": (
        "Creates a real Excel .xlsx spreadsheet FILE from a spoken description "
        "and saves it to the user's Desktop. It can include headers, data rows, "
        "a total row, and a chart. Use for: 'bana bir Excel tablosu oluştur', "
        "'make me a spreadsheet of...', 'haftalık masraflarımı tablo yap ve "
        "toplamını al', 'create an xlsx with monthly sales and a bar chart', "
        "'şu verileri Excel'e dök'. Pass the user's full request verbatim in "
        "'request'. Use this only when the user wants an actual spreadsheet "
        "FILE — not for answering a question about numbers."
    ),
    "parameters": {
        "type": "OBJECT",
        "properties": {
            "request": {
                "type": "STRING",
                "description": "The user's full spreadsheet request, verbatim, in their own language.",
            },
            "filename": {
                "type": "STRING",
                "description": "Optional file name (without extension). If omitted, one is derived from the request.",
            },
        },
        "required": ["request"],
    },
}

_MODEL = "gemini-flash-latest"


# ── output location ───────────────────────────────────────────────────────────

def _output_dir() -> Path:
    desktop = Path.home() / "Desktop"
    return desktop if desktop.is_dir() else Path.home()


def _safe_name(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*]', "", (name or "").strip()) or "spreadsheet"
    return name[:60]


# ── Gemini: request → structured spec ────────────────────────────────────────

def _build_spec(request: str) -> dict:
    from memory.config_manager import get_gemini_key
    api_key = get_gemini_key()
    if not api_key:
        raise RuntimeError("no API key configured")

    from google import genai
    prompt = (
        "Convert the user's request into a spreadsheet specification.\n"
        f'USER REQUEST: "{request}"\n'
        "Return ONLY minified JSON, no markdown fences, with keys:\n"
        ' "title": short sheet title (in the user\'s language),\n'
        ' "filename": short file name without extension (ascii, no spaces→underscores),\n'
        ' "headers": array of column header strings,\n'
        ' "rows": array of rows; each row an array of cell values (numbers as '
        "numbers, text as strings) matching the headers length. Invent "
        "reasonable example data if the user didn't give explicit values,\n"
        ' "total_columns": array of 0-based column indices that should get a '
        "SUM total row (only numeric columns; [] if none),\n"
        ' "chart": either null or {"type":"bar"|"line"|"pie","category_col":int,'
        '"value_col":int,"title":string}.\n'
        "Keep it under 100 rows."
    )
    client = genai.Client(api_key=api_key)
    resp = client.models.generate_content(model=_MODEL, contents=prompt)
    text = (resp.text or "").strip()
    if "{" in text and "}" in text:
        text = text[text.find("{"): text.rfind("}") + 1]
    return json.loads(text)


# ── openpyxl: spec → .xlsx file ───────────────────────────────────────────────

def _build_xlsx(spec: dict, path: Path) -> dict:
    """Writes the workbook. Returns a small summary dict. Raises on hard errors."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers = [str(h) for h in (spec.get("headers") or [])]
    rows = spec.get("rows") or []
    if not headers and rows:
        headers = [f"Column {i+1}" for i in range(len(rows[0]))]
    ncols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = (str(spec.get("title") or "Sheet1"))[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # data rows (pad/trim to header width)
    r = 2
    for row in rows:
        if not isinstance(row, list):
            row = [row]
        for c in range(ncols):
            val = row[c] if c < len(row) else None
            ws.cell(row=r, column=c + 1, value=val)
        r += 1
    last_data_row = r - 1

    # total row (SUM) for requested numeric columns
    total_cols = [i for i in (spec.get("total_columns") or [])
                  if isinstance(i, int) and 0 <= i < ncols]
    if total_cols and last_data_row >= 2:
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        for i in total_cols:
            col = get_column_letter(i + 1)
            cell = ws.cell(row=r, column=i + 1,
                           value=f"=SUM({col}2:{col}{last_data_row})")
            cell.font = Font(bold=True)
        total_row = r
        r += 1
    else:
        total_row = None

    # auto column width
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        width = max([len(str(ws.cell(row=rr, column=c).value or ""))
                     for rr in range(1, last_data_row + 1)] + [8]) + 2
        ws.column_dimensions[col].width = min(40, width)

    # optional chart
    chart_added = False
    chart = spec.get("chart")
    if isinstance(chart, dict) and last_data_row >= 2:
        try:
            ctype = str(chart.get("type", "bar")).lower()
            cat_col = int(chart.get("category_col", 0))
            val_col = int(chart.get("value_col", 1))
            if 0 <= cat_col < ncols and 0 <= val_col < ncols:
                ch = {"line": LineChart, "pie": PieChart}.get(ctype, BarChart)()
                ch.title = chart.get("title") or ws.title
                data_ref = Reference(ws, min_col=val_col + 1, min_row=1, max_row=last_data_row)
                cats_ref = Reference(ws, min_col=cat_col + 1, min_row=2, max_row=last_data_row)
                ch.add_data(data_ref, titles_from_data=True)
                ch.set_categories(cats_ref)
                anchor = f"{get_column_letter(ncols + 2)}2"
                ws.add_chart(ch, anchor)
                chart_added = True
        except Exception:
            chart_added = False

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {"rows": last_data_row - 1, "cols": ncols,
            "total": total_row is not None, "chart": chart_added}


# ── entry point ───────────────────────────────────────────────────────────────

def run(parameters: dict, player=None, session_memory=None) -> str:
    request = (parameters.get("request") or "").strip()
    if not request:
        return "Tell me what the spreadsheet should contain."

    # openpyxl present?
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return ("I need the openpyxl library to build Excel files. "
                "Please install it with: pip install openpyxl")

    try:
        spec = _build_spec(request)
    except Exception as e:
        return f"I couldn't work out the spreadsheet layout: {e}"

    if not (spec.get("headers") or spec.get("rows")):
        return "I couldn't turn that into a table — try describing the columns and data."

    fname = _safe_name(parameters.get("filename") or spec.get("filename") or "spreadsheet")
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    path = _output_dir() / f"{fname}_{stamp}.xlsx"

    try:
        info = _build_xlsx(spec, path)
    except Exception as e:
        return f"I built the layout but couldn't write the file: {e}"

    if player:
        try:
            player.show_content(
                "📊 EXCEL CREATED",
                f"{path.name}\n\n{info['rows']} rows × {info['cols']} columns"
                + ("\n• total row" if info["total"] else "")
                + ("\n• chart" if info["chart"] else "")
                + f"\n\nSaved to:\n{path}")
        except Exception:
            pass

    extras = []
    if info["total"]:
        extras.append("a total row")
    if info["chart"]:
        extras.append("a chart")
    extra_str = f" including {' and '.join(extras)}" if extras else ""
    return (f"Done — I've created {path.name} with {info['rows']} rows{extra_str}, "
            f"and saved it to your Desktop.")
